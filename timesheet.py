#!/usr/bin/env python3
"""Timesheet manager - handles clock-in and clock-out for VS Code extension."""

import os
import sys
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side


def get_timesheet_path(workspace_path):
    """Get the timesheet path in the workspace folder."""
    return os.path.join(workspace_path, "timesheet.xlsx")


def get_base_font():
    """Return the standard font for all cells."""
    return Font(name="Times New Roman", size=12)


def auto_fit_columns(sheet):
    """Auto-fit all columns to their content."""
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Add padding
        sheet.column_dimensions[column_letter].width = adjusted_width


def create_timesheet(timesheet_path):
    """Create a new timesheet with headers."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Timesheet"
    
    # Headers
    headers = ["Date", "Time Clocked-In", "Time Clocked-Out", "Total Duration"]
    
    # Times New Roman, size 12, thick bottom border only
    header_font = get_base_font()
    thick_bottom_border = Border(bottom=Side(style='thick'))
    
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.border = thick_bottom_border
    
    # Auto-fit columns
    auto_fit_columns(sheet)
    
    wb.save(timesheet_path)
    return wb


def get_or_create_workbook(timesheet_path):
    """Load existing timesheet or create new one."""
    if os.path.exists(timesheet_path):
        return load_workbook(timesheet_path)
    return create_timesheet(timesheet_path)


def find_last_data_row(sheet):
    """Find the last row that contains actual timesheet data (has a date in column A)."""
    last_row = 1  # Start after header
    for row in range(2, sheet.max_row + 1):
        date_value = sheet.cell(row=row, column=1).value
        # Only count rows that have a date (Total Duration row won't have a date)
        if date_value is not None:
            last_row = row
    return last_row


def calculate_duration(clock_in_str, clock_out_str):
    """Calculate duration between clock-in and clock-out times in hours."""
    clock_in = datetime.strptime(clock_in_str, "%H:%M:%S")
    clock_out = datetime.strptime(clock_out_str, "%H:%M:%S")
    
    duration = clock_out - clock_in
    hours = duration.total_seconds() / 3600
    return hours


def format_duration(hours):
    """Format duration in hours nicely."""
    # Remove unnecessary trailing zeros
    if hours == int(hours):
        formatted = str(int(hours))
    else:
        formatted = f"{hours:.2f}".rstrip('0').rstrip('.')
    
    if hours == 1:
        return "1 Hour"
    else:
        return f"{formatted} Hours"


def round_to_half_hour(hours):
    """Round hours to nearest half hour."""
    return round(hours * 2) / 2


def calculate_total_duration(sheet, last_data_row):
    """Calculate total duration from all entries."""
    total = 0
    for row in range(2, last_data_row + 1):
        duration_value = sheet.cell(row=row, column=4).value
        if duration_value is not None:
            # Extract the number from "X.X Hours" or "X Hour" format
            try:
                hours_str = str(duration_value).replace(" Hours", "").replace(" Hour", "")
                hours = float(hours_str)
                total += hours
            except:
                pass
    return round_to_half_hour(total)


def update_total_duration_row(sheet, last_data_row):
    """Add or update the Total Duration row."""
    total_row = last_data_row + 1
    base_font = get_base_font()
    
    # Clear any existing total duration rows that might be in wrong positions
    for row in range(2, sheet.max_row + 2):
        if sheet.cell(row=row, column=3).value == "Total Duration:":
            if row != total_row:
                sheet.cell(row=row, column=3).value = None
                sheet.cell(row=row, column=4).value = None
    
    # Add "Total Duration:" label in column C
    label_cell = sheet.cell(row=total_row, column=3, value="Total Duration:")
    label_cell.font = base_font
    
    # Calculate and add total
    total = calculate_total_duration(sheet, last_data_row)
    total_str = format_duration(total)
    total_cell = sheet.cell(row=total_row, column=4, value=total_str)
    total_cell.font = base_font


def clock_action(workspace_path):
    """Perform clock-in or clock-out based on current state."""
    timesheet_path = get_timesheet_path(workspace_path)
    wb = get_or_create_workbook(timesheet_path)
    sheet = wb.active
    
    now = datetime.now()
    current_date = now.strftime("%Y-%m-%d")
    current_time = now.strftime("%H:%M:%S")
    
    base_font = get_base_font()
    
    last_row = find_last_data_row(sheet)
    
    # Check if we need to clock-in or clock-out
    if last_row == 1:
        # No data yet, clock-in
        action = "clock-in"
        new_row = 2
    else:
        clock_out_value = sheet.cell(row=last_row, column=3).value
        if clock_out_value is None or clock_out_value == "":
            # Last row has no clock-out, so clock-out
            action = "clock-out"
            new_row = last_row
        else:
            # Last row is complete, clock-in on new row
            action = "clock-in"
            new_row = last_row + 1
    
    if action == "clock-in":
        # Clear any leftover "Total Duration:" text from this row
        # (This can happen when clocking in on a row that previously held the total)
        sheet.cell(row=new_row, column=3).value = None
        sheet.cell(row=new_row, column=4).value = None
        
        # Add date and clock-in time
        date_cell = sheet.cell(row=new_row, column=1, value=current_date)
        date_cell.font = base_font
        
        clockin_cell = sheet.cell(row=new_row, column=2, value=current_time)
        clockin_cell.font = base_font
        
        print(f"CLOCK-IN: {current_date} at {current_time}")
    else:
        # Add clock-out time
        clockout_cell = sheet.cell(row=new_row, column=3, value=current_time)
        clockout_cell.font = base_font
        
        # Calculate duration
        clock_in_time = sheet.cell(row=new_row, column=2).value
        duration = calculate_duration(clock_in_time, current_time)
        duration_str = format_duration(duration)
        
        duration_cell = sheet.cell(row=new_row, column=4, value=duration_str)
        duration_cell.font = base_font
        
        # Update total duration row
        update_total_duration_row(sheet, new_row)
        
        print(f"CLOCK-OUT: {current_date} at {current_time}")
        print(f"Duration: {duration_str}")
    
    # Auto-fit columns
    auto_fit_columns(sheet)
    
    wb.save(timesheet_path)
    print(f"Timesheet saved to: {timesheet_path}")
    return action


if __name__ == "__main__":
    try:
        if len(sys.argv) < 2:
            print("ERROR: Workspace path not provided", file=sys.stderr)
            sys.exit(1)
        
        workspace_path = sys.argv[1]
        
        if not os.path.isdir(workspace_path):
            print(f"ERROR: Invalid workspace path: {workspace_path}", file=sys.stderr)
            sys.exit(1)
        
        action = clock_action(workspace_path)
        sys.exit(0)
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)